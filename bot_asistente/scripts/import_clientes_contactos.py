"""
import_clientes_contactos.py — Carga contactos a la tabla `clientes`.

A diferencia de `import_contactos_vcf.py` (que carga a `numeros_internos`
para que Laura los IGNORE), este script carga los contactos como CLIENTES
con nombre, para que cuando ellos escriban Laura ya los conozca por nombre.

Soporta:
  - CSV con cabeceras (numero, nombre) en cualquier orden, con cualquier
    separador común (',' ';' '\\t').
  - .vcf (vCards exportadas de WhatsApp)
  - .xlsx (Excel)

Idempotente: si el cliente ya existe (por número), solo actualiza el
nombre si está vacío. Nunca sobreescribe un nombre ya guardado.

Uso:
    python scripts/import_clientes_contactos.py --archivo ~/clientes.csv
    python scripts/import_clientes_contactos.py --archivo contactos.vcf --dry-run
    python scripts/import_clientes_contactos.py --archivo lista.xlsx
"""

from __future__ import annotations

import argparse
import csv
import os
import re
import sys
from pathlib import Path

import psycopg2  # type: ignore
from psycopg2.extras import execute_values  # type: ignore


DEFAULT_DSN = os.getenv(
    "DATABASE_URL_SYNC",
    "postgresql://asistente_user:Colombia1234.@127.0.0.1:5432/asistente_db",
)

RE_TEL_VCF = re.compile(r"TEL[^:]*:(.+)", re.IGNORECASE)
RE_FN_VCF = re.compile(r"^FN:(.+)$", re.IGNORECASE)
RE_WAID = re.compile(r"waid=(\d+)", re.IGNORECASE)
RE_DIGITS = re.compile(r"[^\d+]")


def normalizar_numero(raw: str) -> str | None:
    """Devuelve número en formato E.164 +57XXXXXXXXXX. None si inválido."""
    if not raw:
        return None
    s = RE_DIGITS.sub("", str(raw).strip())
    if not s:
        return None
    if s.startswith("+"):
        s = s[1:]
    # Quita ceros iniciales
    while s.startswith("0"):
        s = s[1:]
    # Casos Colombia
    if len(s) == 10 and s.startswith("3"):  # 3001234567 → +573001234567
        return f"+57{s}"
    if len(s) == 12 and s.startswith("57"):  # 573001234567 → +573001234567
        return f"+{s}"
    if len(s) >= 10:
        return f"+{s}"
    return None


def limpiar_nombre(raw: str) -> str:
    """Trim, colapsa espacios, recorta a 255 chars (límite columna)."""
    if not raw:
        return ""
    s = " ".join(str(raw).split())
    return s[:255]


def parsear_csv(path: Path) -> list[tuple[str, str]]:
    """Lee CSV, retorna lista de (numero_e164, nombre)."""
    contactos: list[tuple[str, str]] = []
    # Sniff del delimitador
    with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
        muestra = f.read(4096)
        f.seek(0)
        try:
            dialecto = csv.Sniffer().sniff(muestra, delimiters=",;\t|")
        except csv.Error:
            dialecto = csv.excel
        reader = csv.reader(f, dialect=dialecto)
        rows = list(reader)
    if not rows:
        return []

    # Detectar header
    primera = [c.lower().strip() for c in rows[0]]
    header_keywords_num = {"numero", "número", "telefono", "teléfono", "phone", "tel", "celular", "whatsapp"}
    header_keywords_name = {"nombre", "name", "contacto", "cliente"}
    idx_num, idx_nom = None, None
    if any(k in c for c in primera for k in header_keywords_num) or any(k in c for c in primera for k in header_keywords_name):
        for i, col in enumerate(primera):
            if idx_num is None and any(k in col for k in header_keywords_num):
                idx_num = i
            if idx_nom is None and any(k in col for k in header_keywords_name):
                idx_nom = i
        rows = rows[1:]  # quitar header

    # Si no hubo header, asumir col 0 = número, col 1 = nombre
    if idx_num is None:
        idx_num = 0
    if idx_nom is None:
        idx_nom = 1 if any(len(r) > 1 for r in rows) else 0

    for r in rows:
        if not r:
            continue
        num_raw = r[idx_num] if idx_num < len(r) else ""
        nom_raw = r[idx_nom] if idx_nom < len(r) else ""
        num = normalizar_numero(num_raw)
        nom = limpiar_nombre(nom_raw)
        if num:
            contactos.append((num, nom))
    return contactos


def parsear_vcf(path: Path) -> list[tuple[str, str]]:
    """Lee .vcf (multi-vCard), retorna lista de (numero, nombre)."""
    contactos: list[tuple[str, str]] = []
    contenido = path.read_text(encoding="utf-8", errors="replace")
    # Cada vCard delimitada por BEGIN:VCARD ... END:VCARD
    bloques = re.split(r"BEGIN:VCARD", contenido, flags=re.IGNORECASE)[1:]
    for b in bloques:
        nombre = ""
        numero = None
        for linea in b.splitlines():
            linea = linea.strip()
            if m := RE_FN_VCF.match(linea):
                nombre = limpiar_nombre(m.group(1))
            elif linea.upper().startswith("TEL"):
                waid_m = RE_WAID.search(linea)
                if waid_m:
                    numero = f"+{waid_m.group(1)}"
                elif ":" in linea:
                    numero = normalizar_numero(linea.split(":", 1)[1])
        if numero:
            contactos.append((numero, nombre))
    return contactos


def parsear_xlsx(path: Path) -> list[tuple[str, str]]:
    """Lee .xlsx — asume col A = numero, col B = nombre. Skip header si lo detecta."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        print("ERROR: para leer .xlsx instala openpyxl: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    contactos: list[tuple[str, str]] = []
    primera = True
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        a, b = (row[0] if len(row) > 0 else None), (row[1] if len(row) > 1 else None)
        if primera:
            primera = False
            # Si la primera fila parece header, skip
            txt = " ".join(str(x or "").lower() for x in row[:2])
            if any(k in txt for k in ("nombre", "numero", "tel", "phone", "name")):
                continue
        num = normalizar_numero(a)
        nom = limpiar_nombre(b)
        if num:
            contactos.append((num, nom))
    return contactos


def deduplicar(contactos: list[tuple[str, str]]) -> list[tuple[str, str]]:
    """Quita duplicados por número. Mantiene la primera ocurrencia."""
    visto: dict[str, str] = {}
    for num, nom in contactos:
        if num not in visto or (not visto[num] and nom):
            visto[num] = nom
    return [(n, nom) for n, nom in visto.items()]


def upsert(contactos: list[tuple[str, str]], dsn: str, dry_run: bool = False) -> dict:
    """Inserta o actualiza en `clientes`. Política: no sobreescribir nombre existente."""
    if dry_run:
        return {"insertados": 0, "actualizados": 0, "skipped_sin_cambio": len(contactos)}

    insertados = 0
    actualizados = 0
    sin_cambio = 0

    conn = psycopg2.connect(dsn)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            # Batch upsert: insertar si no existe; si existe pero nombre vacío, completar.
            execute_values(
                cur,
                """
                INSERT INTO clientes (numero_whatsapp, nombre, primer_contacto, ultimo_contacto)
                VALUES %s
                ON CONFLICT (numero_whatsapp) DO UPDATE
                  SET nombre = COALESCE(NULLIF(clientes.nombre, ''), EXCLUDED.nombre)
                  WHERE COALESCE(NULLIF(clientes.nombre, ''), '') = ''
                RETURNING (xmax = 0) AS inserted, id
                """,
                [(n, nom or None, "now()", "now()") for n, nom in contactos],
                template="(%s, %s, now(), now())",
            )
            for inserted, _id in cur.fetchall():
                if inserted:
                    insertados += 1
                else:
                    actualizados += 1
            sin_cambio = len(contactos) - insertados - actualizados
        conn.commit()
    finally:
        conn.close()
    return {"insertados": insertados, "actualizados": actualizados, "skipped_sin_cambio": sin_cambio}


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--archivo", required=True, help="Ruta al CSV/.vcf/.xlsx")
    ap.add_argument("--dry-run", action="store_true", help="Solo parsear y reportar, sin escribir BD")
    ap.add_argument("--dsn", default=DEFAULT_DSN)
    args = ap.parse_args()

    path = Path(args.archivo).expanduser().resolve()
    if not path.exists():
        print(f"ERROR: no existe {path}", file=sys.stderr)
        return 1

    ext = path.suffix.lower()
    if ext == ".csv":
        contactos = parsear_csv(path)
    elif ext == ".vcf":
        contactos = parsear_vcf(path)
    elif ext in (".xlsx", ".xlsm"):
        contactos = parsear_xlsx(path)
    else:
        print(f"ERROR: extensión no soportada: {ext}. Usa .csv, .vcf o .xlsx", file=sys.stderr)
        return 1

    print(f"📋 Leídos {len(contactos)} contactos de {path.name}")
    contactos = deduplicar(contactos)
    print(f"   {len(contactos)} únicos por número")
    con_nombre = sum(1 for _, n in contactos if n)
    sin_nombre = len(contactos) - con_nombre
    print(f"   con nombre: {con_nombre} · sin nombre: {sin_nombre}")

    if args.dry_run:
        print("\n🧪 Dry-run. Primeros 5:")
        for n, nom in contactos[:5]:
            print(f"   {n} → {nom or '(sin nombre)'}")
        return 0

    print("\n📤 Upsert a Postgres …")
    res = upsert(contactos, args.dsn, dry_run=False)
    print(f"✅ insertados nuevos: {res['insertados']}")
    print(f"   actualizado nombre (estaba vacío): {res['actualizados']}")
    print(f"   skipped (ya existían con nombre): {res['skipped_sin_cambio']}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
